import openpyxl
from spellchecker import SpellChecker
import language_tool_python
from random_string_detector import RandomStringDetector

spell_checker = SpellChecker(distance=4)
invalid_name = RandomStringDetector(allow_numbers=True)
grammar_checker = language_tool_python.LanguageTool('en-US')

def read_xlsx(path, sheet_num):
    accounts = {}
    workbook = openpyxl.load_workbook(path)

    # if the number for the sheet in the workbook is invalid change the number to the last sheet in the page
    sheet_num = sheet_num if sheet_num <= len(workbook.sheetnames) else len(workbook.sheetnames)
    sheet = workbook[workbook.sheetnames[sheet_num - 1]]

    # number of rows and columns in the sheet
    row_l = sheet.max_row
    col_l = sheet.max_column

    for row_i in range(2, row_l + 1): # through the list starting from 2nd row in the sheet
        post = {}

        user = sheet.cell(row = row_i, column = 4).value # get the username of the account
        accounts[user] = [] if user not in accounts else accounts[user] # add user to the dictionary of accounts

        for col_j in range(1, col_l + 1):
            key = sheet.cell(row = 1, column = col_j).value
            val = sheet.cell(row = row_i, column = col_j).value
            post[key] = val

        accounts[user].append(post) # add the post to the list of posts for each user

    return accounts

def find_bots(accounts):
    scrmbl_nms, freq_pstr, mny_mstk = set(), set(), set()
    for acc in accounts:
        user = accounts[acc]
        username = (user[0])['author'] # grab the username so we can add it to the set if necessary
        # checks if username is alphanumerically scrambled:
        if username not in scrmbl_nms and invalid_name(username):
            scrmbl_nms.add(username)
        # checks if there are a lot of spelling or grammatical mistakes in the post:
        if username not in mny_mstk and many_mistakes_in_text(user):
            mny_mstk.add(username)
        # checks if the user posts frequently:
        if username not in freq_pstr and posts_frequently(user):
           freq_pstr.add(username)
    return {'scrambled names': scrmbl_nms, 'frequent posters': freq_pstr, 'many mistakes': mny_mstk}

def posts_frequently(user_posts):
    hours, days = [], []
    for post in user_posts:
        days.append((post['date'].split('T'))[0]) # days
        hours.append((post['date'].split(':'))[0]) # hours of the day
    # if someone posts at least a dozen times per day or at least 4 times in an hour, they may be a bot
    return True if any(days.count(d) >= 12 for d in days) or any(hours.count(h) >= 4 for h in hours) else False

def many_mistakes_in_text(posts):
    for post in posts:
        word_list = []
        text = post['text']
        lang = post['language']

        # remove any "words" that may contain numbers or symbols as they increase the number of false positives
        for word in text.split():
            word_list.append(word if not any(not char.isalpha() for char in word) else None)
        word_list = list(filter(None, word_list))

        # if the language is english
        if lang is not None and lang[:2] == 'en':
            # check for half a dozen spelling or grammatical errors
            if len(spell_checker.unknown(word_list)) >= 6 or len(grammar_checker.check(text)) >= 6:
                return True
    return False

def main():
    path = input("Please provide the path to the file: ") # please use the absolute path of the file
    sheet = int(input("Please provide the number of the sheet: "))

    account_list = read_xlsx(path, sheet)
    bots = find_bots(account_list)
    grammar_checker.close() # closes the grammar checker after use

    for category in bots:
        print('\n----Potential bots based on looking for:', category)
        for bot in bots[category]: print(bot)

    total_bots = len(bots['scrambled names'] | bots['frequent posters'] | bots['many mistakes']) # combines the lists
    print("\n----Total number of potential bots:", total_bots)

main()
